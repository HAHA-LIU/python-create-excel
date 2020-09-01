# -*- coding:utf-8 -*-
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl.styles import Alignment,Font
from greentransport.settings import BASE_DIR


def create_excel(excel_name=None,data=None,file_result=None,fun_result=None):
	# 创建excel文件路径
	filename = excel_name
	path_gen = BASE_DIR + '/'
	path_file = 'static/excel/' + filename + '.xlsx'
	path = path_gen + path_file
	print('path:',path)

	workbook = Workbook()

	# 激活表格
	sheet = workbook.active

	# 合并单元格
	sheet.merge_cells('A1:F1')	# 大标题 表 城市绿色货运配送示范工程运行情况信息表
	sheet.merge_cells('A2')		# 小标题 示范城市：
	sheet.merge_cells('B2')		# 小标题 示范期：
	sheet.merge_cells('C2')		# 小标题 填报人：
	sheet.merge_cells('D2')		# 小标题 联系电话：
	sheet.merge_cells('E2')		# 小标题 填报周期：__年第__季度
	sheet.merge_cells('A3')		# 小标题 指标分类：

	# 小标题 具体指标：
	for i in range(3,36):
		ceil = '{}:{}'.format('B' + str(i), 'C' + str(i))
		sheet.merge_cells(ceil)

	# 小标题 值与值的空
	for i in range(3,36):
		ceil = '{}'.format('D'+str(i))
		sheet.merge_cells(ceil)

	sheet.merge_cells('A4:A11')		# 指标分类 体制机制保障
	sheet.merge_cells('A12:A14')	# 指标分类 城市配送物流基础设施
	sheet.merge_cells('A15:A22')	# 指标分类 城市配送车辆及配套设施
	sheet.merge_cells('A23:A28')	# 指标分类 便利通行政策
	sheet.merge_cells('A29:A30')	# 指标分类 先进配送组织模式
	sheet.merge_cells('A31')		# 指标分类 信息化建设
	sheet.merge_cells('A32')		# 指标分类 市场主体培育
	sheet.merge_cells('A33:A34')	# 指标分类 物流降本增效
	sheet.merge_cells('A35')		# 指标分类 节能减排

	# 单元格样式
	sheet['A1'].alignment = Alignment(horizontal='center',vertical='center')  	# 水平居中,上下居中
	sheet['A2'].alignment = Alignment(horizontal='left',vertical='center')		# 示范城市
	sheet['B2'].alignment = Alignment(horizontal='left',vertical='center')		# 示范期
	sheet['C2'].alignment = Alignment(horizontal='left',vertical='center')		# 填报人
	sheet['D2'].alignment = Alignment(horizontal='left',vertical='center')		# 联系电话
	sheet['E2'].alignment = Alignment(horizontal='left',vertical='center')		# 填报周期

	sheet['A1'].font = Font(bold=True, name='宋体', size=12 ) 	# 加粗，宋体，12
	sheet['A2'].font = Font(name='宋体', size=12 )				# 示范城市
	sheet['B2'].font = Font(name='宋体', size=12 )				# 示范期
	sheet['C2'].font = Font(name='宋体', size=12 )				# 填报人
	sheet['D2'].font = Font(name='宋体', size=12 )				# 联系电话
	sheet['E2'].font = Font(name='宋体', size=12 )				# 填报周期

	sheet['A3'].font = Font(bold=True, name='仿宋', size=12 )	# 指标分类
	sheet['B3'].font = Font(bold=True, name='仿宋', size=12 )	# 具体指标
	sheet['D3'].font = Font(bold=True, name='仿宋', size=12 )	# 值
	sheet['E3'].font = Font(bold=True, name='仿宋', size=12 )	# 描述
	sheet['F3'].font = Font(bold=True, name='仿宋', size=12 )	# 附件名称

	sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
	sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
	sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
	sheet['E3'].alignment = Alignment(horizontal='center', vertical='center')
	sheet['F3'].alignment = Alignment(horizontal='center', vertical='center')

	sheet['A4'].alignment = Alignment(horizontal='center',vertical='center')	# 体制机制保障
	sheet['A12'].alignment = Alignment(horizontal='center',vertical='center')	# 城市配送物流基础设施
	sheet['A15'].alignment = Alignment(horizontal='center',vertical='center')	# 城市配送车辆及配套设施
	sheet['A23'].alignment = Alignment(horizontal='center',vertical='center')	# 便利通行政策
	sheet['A29'].alignment = Alignment(horizontal='center',vertical='center')	# 先进配送组织模式
	sheet['A31'].alignment = Alignment(horizontal='center',vertical='center')	# 信息化建设
	sheet['A32'].alignment = Alignment(horizontal='center',vertical='center')	# 市场主体培育
	sheet['A33'].alignment = Alignment(horizontal='center',vertical='center')	# 物流降本增效
	sheet['A35'].alignment = Alignment(horizontal='center',vertical='center')	# 节能减排

	sheet['A4'].font = Font(name='仿宋', size=11 )
	sheet['A12'].font = Font(name='仿宋', size=11 )
	sheet['A15'].font = Font(name='仿宋', size=11 )
	sheet['A23'].font = Font(name='仿宋', size=11 )
	sheet['A29'].font = Font(name='仿宋', size=11 )
	sheet['A31'].font = Font(name='仿宋', size=11 )
	sheet['A32'].font = Font(name='仿宋', size=11 )
	sheet['A33'].font = Font(name='仿宋', size=11 )
	sheet['A35'].font = Font(name='仿宋', size=11 )

	# 行高 列宽
	for i in range(1,4):
		sheet.row_dimensions[i].height = 30	# 行高 1-3行
	for i in range(4,36):
		sheet.row_dimensions[i].height = 26	# 行高 4-32行

	sheet.column_dimensions['A'].width = 25	# 列宽
	sheet.column_dimensions['B'].width = 42	# 列宽
	sheet.column_dimensions['C'].width = 24	# 列宽
	sheet.column_dimensions['D'].width = 30	# 列宽
	sheet.column_dimensions['E'].width = 34	# 列宽
	sheet.column_dimensions['F'].width = 34	# 列宽

	# 给题目特殊的单元格赋值
	sheet['A1'] = '表 城市绿色货运配送示范工程运行情况信息表'
	sheet['A2'] = '示范城市：' + data[0][1]
	sheet['B2'] = '示范期：' + data[0][2] + ' 至 ' + data[0][3]
	sheet['C2'] = '填报人：' + data[0][4]
	sheet['D2'] = '联系电话：' + data[0][5]
	sheet['E2'] = '填报周期：' + str(data[0][6]) + '年第' + str(data[0][7]) + '季度'
	sheet['A3'] = '指标分类：'
	sheet['B3'] = '具体指标：'
	sheet['D3'] = '值：'
	sheet['E3'] = '描述:'
	sheet['F3'] = '附件名称:'

	sheet['A4'] = '体制机制保障'
	sheet['A12'] = '城市配送物流基础设施'
	sheet['A15'] = '城市配送车辆及配套设施'
	sheet['A23'] = '便利通行政策'
	sheet['A29'] = '先进配送组织模式'
	sheet['A31'] = '信息化建设'
	sheet['A32'] = '市场主体培育'
	sheet['A33'] = '物流降本增效'
	sheet['A35'] = '节能减排'


	# 问题 答案 附加描述 附件名 赋值 len(data)+4 是多少行  data[i-4] data里面第一个元素
	for i in range(4,len(data)+4):
		ceil_B = 'B' + str(i)			# 问题单元格坐标
		sheet[ceil_B] = data[i-4][14]  	# 问题赋值
		sheet[ceil_B].alignment = Alignment(horizontal='left', vertical='center')  # 左对齐，垂直居中
		sheet[ceil_B].font = Font(name='仿宋', size=11)

		ceil_D = 'D' + str(i)			# 答案单元格坐标
		# 首先判断问题类型type 索引13，value 索引10
		# 1：是否
		if data[i-4][13] == 1:
			if data[i-4][10] == '1':
				sheet[ceil_D] = '是'
			elif data[i-4][10] == '2':
				sheet[ceil_D] = '否'

		# 2：文本框  文本框数据类型  1：整数  2：两位小数  3：百分比  4：只读
		elif data[i-4][13] == 2:
			if data[i-4][16] == 3:
				sheet[ceil_D] = data[i-4][10] + '%'
			else:
				sheet[ceil_D] = data[i-4][10]

		# 3：有无(是否需要上传附件)
		elif data[i-4][13] == 3:
			# 需要上传附件
			if data[i-4][17] == 1:
				# 判断值
				if data[i-4][10] == '1':
					sheet[ceil_D] = '有'
					sheet['E' + str(i)] = data[i-4][12]  # 描述
					# 将附件名称备注
					mark = ''
					for item in file_result: # [(('20', '我是20题附件1'),), (('24', '我是24题附件1'),)]
						for tar in item:
							if tar[0] == data[i-4][-1]:
								mark += tar[1] + ';'
					if not mark:
						sheet['F' + str(i)] = '未上传附件'
					else:
						sheet['F' + str(i)] = mark
				elif data[i-4][10] == '2':
					sheet[ceil_D] = '无'
					sheet['F' + str(i)] = '未上传附件'
					sheet['E' + str(i)] = data[i-4][12]  # 描述

			# 不需要上传附件
			elif data[i-4][17] == 2:
				# 判断值
				if data[i - 4][10] == '1':
					sheet[ceil_D] = '有'
					sheet['E' + str(i)] = data[i - 4][12]  # 描述

				elif data[i - 4][10] == '2':
					sheet[ceil_D] = '无'
					sheet['E' + str(i)] = data[i - 4][12]  # 描述

		# 4：是否（附带复选框和说明）
		elif data[i-4][13] == 4:
			str_q = ''
			if data[i-4][10] == '1':
				sheet[ceil_D] = '有'
				sheet['E'+str(i)] = data[i-4][12]
				# 对功能进行匹配
				for item in list(eval(data[i-4][11])):
					for j in fun_result:
						if str(item) == j[0]:
							str_q += j[1] + ' '
				sheet['E'+str(i)] = data[i-4][12] + ' ' + str_q
			elif data[i-4][10] == '2':
				sheet[ceil_D] = '无'
		sheet[ceil_D].alignment = Alignment(horizontal='center', vertical='center')  # 水平居中，垂直居中
		sheet[ceil_D].font = Font(name='仿宋', size=11)

	# 临时文件 将表格数据保存到流中
	#with NamedTemporaryFile('w+b',delete=False) as tmp:				# windows
	# with NamedTemporaryFile('w+b') as tmp:						# linux
		#print('Tmp:',tmp.name)
		#workbook.save(tmp.name)
		#tmp.seek(0)
		#stream = tmp.read()
	#print('表格生成完成')
	#return stream

	# 保留文件
	workbook.save(filename=path)
	print('表格生成完成')
	return path_file

if __name__ == '__main__':
	# 运行这里，放开保留文件那块代码，注释临时文件代码
	data = (('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '1', '2', None, None, 1, '创建领导小组', '1', None, None, '体制机制保障', '1'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '2', '12', None, None, 2, '协同推进工作（次）', '1', 1, None, '体制机制保障', '2'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '3', '2', None, None, 1, '制定城市配送企业考核管理办法', '1', None, None, '体制机制保障', '3'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '4', '2', None, None, 1, '建立快递车辆规范管理制度', '1', None, None, '体制机制保障', '4'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '5', '13234.34', None, None, 2, '新能源车辆购置与营运方面政府拨付资金（元）', '1', 2, None, '体制机制保障', '5'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '6', '1234.09', None, None, 2, '配送中心建设方面政府拨付资金（元）', '1', 2, None, '体制机制保障', '6'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '7', '7567.75', None, None, 2, '先进组织模式推广应用方面政府拨付资金（元）', '1', 2, None, '体制机制保障', '7'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '8', '142345.86', None, None, 2, '市场主体培育方面政府拨付资金（元）', '1', 2, None, '体制机制保障', '8'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '10', '13', None, None, 2, '公共配送中心（个）', '2', 4, None, '城市配送物流基础设施', '10'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '11', '12', None, None, 2, '末端共同配送站（个）', '2', 1, None, '城市配送物流基础设施', '11'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '9', '2', None, None, 2, '干支衔接型货运枢纽（物流园区）（个）', '2', 4, None, '城市配送物流基础设施', '9'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '12', '11', None, None, 2, '新增纯电动城市配送营运车辆（辆）', '3', 1, None, '城市配送车辆及配套设施', '12'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '13', '22', None, None, 2, '新增新能源城市配送营运车辆（辆）', '3', 1, None, '城市配送车辆及配套设施', '13'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '14', '33', None, None, 2, '新增和更新城市配送营运车辆（辆）', '3', 1, None, '城市配送车辆及配套设施', '14'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '15', '44', None, None, 2, '示范建设期初城市配送新能源纯电动货车及插电式混合动力货车保有量（辆）', '3', 1, None, '城市配送车辆及配套设施', '15'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '16', '55', None, None, 2, '当期城市配送新能源纯电动货车及插电式混合动力货车保有量（辆）', '3', 1, None, '城市配送车辆及配套设施', '16'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '17', '66', None, None, 2, '新能源物流配送车辆充电桩数量（个）', '3', 1, None, '城市配送车辆及配套设施', '17'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '18', '88', None, None, 2, '冷藏保温配送车辆保有量（辆）', '3', 1, None, '城市配送车辆及配套设施', '18'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '19', '77', None, None, 2, '城市配送车辆保有量（辆）', '3', 1, None, '城市配送车辆及配套设施', '19'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '20', '1', None, '我有新能源物流配送车辆综合性政策', 3, '新能源物流配送车辆综合性政策', '4', None, 1, '便利通行政策', '20'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '21', '2', None, None, 3, '新能源物流配送车辆通行政策', '4', None, 1, '便利通行政策', '21'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '22', '2', None, None, 3, '新能源物流配送车辆停车便利政策', '4', None, 1, '便利通行政策', '22'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '23', '2', None, None, 3, '城市货运配送需求调查预测制度', '4', None, 2, '便利通行政策', '23'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '24', '1', None, '我有配送车辆分时、错时、分类通行和停放措施', 3, '配送车辆分时、错时、分类通行和停放措施', '4', None, 1, '便利通行政策', '24'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '25', '18', None, None, 2, '城市货运配送车辆临时停靠点数量（个）', '4', 1, None, '便利通行政策', '25'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '26', '12.43', None, None, 2, '中心区大型超市（卖场）、连锁店等商贸流通企业采用共同（集中）配送的比例（%）', '5', 3, None, '先进配送组织模式', '26'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '27', '35.67', None, None, 2, '中心区大型超市（卖场）、连锁店等商贸流通企业采用夜间配送的比例（%）', '5', 3, None, '先进配送组织模式', '27'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '28', '1', '1,2,4', '我的信息化平台实现了政务发布，车辆监管等功能', 4, '城市货运配送公共信息服务平台', '6', None, None, '信息化建设', '28'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '29', '4', None, None, 2, 'AAA级（含）以上城市配送企业数量（个）', '7', 1, None, '市场主体培育', '29'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '30', '22.32', None, None, 2, '城市配送成本较示范建设期初降低的比例', '8', 3, None, '物流降本增效', '30'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '31', '35.34', None, None, 2, '城市配送车辆利用效率较示范建设期初提高的比例', '8', 3, None, '物流降本增效', '31'), ('1', '大同市', '2020-08-17', '2020-08-27', '小明', '13533484483', 2020, 1, '1', '32', '44.23', None, None, 2, '城市配送车辆单位周转量能耗较示范建设期初降低的比例', '9', 3, None, '节能常量', '32'))
	file_result = [(('20', '我是20题附件1'), ('20', '我是20题附件2')), (('24', '我是24题附件1'), ('24', '我是24题附件2'))]
	fun_result = (('1', '政务发布'), ('2', '信息查询'), ('3', '车辆监管'), ('4', '数据分析'), ('5', '电子围栏'), ('6', '交通诱导'))
	create_excel(excel_name='数据详情',data=data,file_result=file_result,fun_result=fun_result)
